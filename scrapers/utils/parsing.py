"""
parsing.py — Shared parsing utilities for HTML, PDF, and spreadsheet sources.

These helpers are used by utility-specific scrapers so common logic
(extracting tables from HTML, reading PDF tables, cleaning numeric
strings) lives in one place.
"""

from __future__ import annotations

import io
import re
import csv
import hashlib
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Iterable, Optional
from urllib.parse import urljoin, urlsplit

from bs4 import BeautifulSoup

logger = logging.getLogger(__name__)


@dataclass(frozen=True)
class DocumentPage:
    """Text extracted from one source page, retaining its human page number."""

    page_number: int
    text: str

    @property
    def source_detail(self) -> str:
        return f"PDF page {self.page_number}"


# ─── HTML helpers ──────────────────────────────────────────────

def parse_html(html: str) -> BeautifulSoup:
    """Parse HTML into a BeautifulSoup tree using the fast lxml parser."""
    return BeautifulSoup(html, "lxml")


def extract_tables(html: str) -> list[list[list[str]]]:
    """
    Extract all <table> elements from HTML and return them as
    a list of tables, where each table is a list of rows, and
    each row is a list of cell text strings.

    Example return:
        [
            [                          # table 0
                ["Header 1", "Header 2"],
                ["Value 1",  "Value 2"],
            ],
            [                          # table 1
                ...
            ],
        ]
    """
    soup = parse_html(html)
    tables = []
    for table_tag in soup.find_all("table"):
        rows = []
        for tr in table_tag.find_all("tr"):
            cells = []
            for cell in tr.find_all(["td", "th"]):
                text = cell.get_text(strip=True)
                cells.append(text)
            if cells:
                rows.append(cells)
        if rows:
            tables.append(rows)
    return tables


def find_table_by_header(html: str, header_text: str) -> Optional[list[list[str]]]:
    """
    Find the first table whose first row contains header_text
    (case-insensitive substring match).
    """
    tables = extract_tables(html)
    header_lower = header_text.lower()
    for table in tables:
        if table and any(header_lower in cell.lower() for cell in table[0]):
            return table
    return None


# ─── PDF helpers ───────────────────────────────────────────────

def extract_pdf_tables(pdf_bytes: bytes) -> list[list[list[str]]]:
    """
    Extract tables from a PDF using pdfplumber.

    Returns the same format as extract_tables():
    list of tables, each a list of rows, each a list of cell strings.

    Requires:  pip install pdfplumber
    """
    try:
        import pdfplumber
    except ImportError:
        logger.error("pdfplumber not installed — run: pip install pdfplumber")
        return []

    tables = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_tables = page.extract_tables()
            if page_tables:
                for table in page_tables:
                    cleaned = []
                    for row in table:
                        cleaned_row = [cell.strip() if cell else "" for cell in row]
                        cleaned.append(cleaned_row)
                    tables.append(cleaned)
    return tables


def extract_pdf_text(pdf_bytes: bytes) -> str:
    """Extract raw text from all pages of a PDF."""
    try:
        import pdfplumber
    except ImportError:
        logger.error("pdfplumber not installed — run: pip install pdfplumber")
        return ""

    text_parts = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
    return "\n\n".join(text_parts)


def extract_pdf_pages(pdf_bytes: bytes, minimum_characters: int = 20) -> list[DocumentPage]:
    """Extract usable PDF pages and fail closed on empty/corrupt documents.

    Page boundaries are deliberately retained so a component can cite an exact
    location. Repeated whitespace is normalized, while minus signs and
    parentheses (which commonly identify credits) are preserved.
    """
    try:
        import pdfplumber
    except ImportError:
        logger.error("pdfplumber not installed — run: pip install pdfplumber")
        return []

    pages: list[DocumentPage] = []
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for number, page in enumerate(pdf.pages, 1):
                raw = page.extract_text(x_tolerance=2, y_tolerance=3) or ""
                text = normalize_document_text(raw)
                if len(re.sub(r"\W", "", text)) >= minimum_characters:
                    pages.append(DocumentPage(number, text))
    except Exception as exc:
        logger.warning("PDF extraction failed closed: %s", exc)
        return []
    return pages


def normalize_document_text(text: str) -> str:
    """Normalize extracted document text and reconstruct wrapped rows."""
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    repeated = {line for line in lines if line and lines.count(line) > 2}
    output: list[str] = []
    for line in lines:
        if not line or line in repeated:
            continue
        # A line with no numeric value is normally a wrapped label. Keep it on
        # the same logical row as the following value-bearing line.
        if output and not re.search(r"\d", output[-1]) and re.search(r"\d", line):
            output[-1] = f"{output[-1]} {line}"
        else:
            output.append(line)
    return "\n".join(output)


def find_pdf_section(
    pages: Iterable[DocumentPage],
    section_labels: Iterable[str],
    *,
    tariff_code: Optional[str] = None,
    following_pages: int = 1,
) -> list[DocumentPage]:
    """Return pages belonging to a labelled tariff section.

    Merely finding a number is insufficient: at least one requested label and,
    when supplied, the tariff code must occur on the anchor page. A limited
    number of following pages supports schedules split across page boundaries.
    """
    page_list = list(pages)
    labels = [label.casefold() for label in section_labels]
    for index, page in enumerate(page_list):
        folded = page.text.casefold()
        if not any(label in folded for label in labels):
            continue
        if tariff_code and not re.search(rf"\b{re.escape(tariff_code)}\b", page.text, re.I):
            continue
        return page_list[index:index + following_pages + 1]
    return []


# ─── Spreadsheet helpers ──────────────────────────────────────

def read_xlsx_tables(xlsx_bytes: bytes) -> dict[str, list[list[str]]]:
    """
    Read an Excel file and return a dict of sheet_name -> rows.
    Each row is a list of cell values as strings.

    Requires:  pip install openpyxl
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed — run: pip install openpyxl")
        return {}

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    result = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        result[sheet_name] = rows
    return result


def read_csv_rows(csv_bytes: bytes) -> list[dict[str, str]]:
    """Read an official UTF-8/UTF-8-BOM CSV into named rows."""
    text = csv_bytes.decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(text)))


def content_hash(content: bytes | str) -> str:
    """Return a stable SHA-256 hash for downloaded official content."""
    data = content.encode("utf-8") if isinstance(content, str) else content
    return hashlib.sha256(data).hexdigest()


# ─── Numeric / text cleaning ─────────────────────────────────

def clean_currency(text: str) -> Optional[float]:
    """
    Parse a currency string like "$0.0945", "12.34¢/kWh", "0.0945 $/kWh"
    into a float (always in dollars, not cents).

    Returns None if parsing fails.
    """
    if not text:
        return None

    text = text.strip()

    # Remove common non-numeric decorations
    text = text.replace(",", "")
    text = text.replace(" ", "")

    # Check for cents
    is_cents = "¢" in text or "cents" in text.lower()

    # Parentheses and a leading minus are both published credit conventions.
    negative = text.startswith("-") or bool(re.match(r"^\(\d", text))
    match = re.search(r"(\d+\.?\d*)", text)
    if not match:
        return None

    value = float(match.group(1))

    # Convert cents to dollars
    if is_cents:
        value = value / 100.0

    return -value if negative else value


def normalize_charge_unit(text: str) -> Optional[str]:
    """Normalize Canadian tariff units without conflating kW and kVA."""
    value = text.casefold().replace("³", "3").replace("cubic metre", "m3")
    period = "month" if re.search(r"month|monthly", value) else "day" if re.search(r"day|daily", value) else None
    measure = next((unit for token, unit in (
        ("kwh", "kWh"), ("kva", "kVA"), ("kw", "kW"),
        ("gj", "GJ"), ("m3", "m3"),
    ) if token in value), None)
    denominator = measure or period
    if not denominator:
        return None
    return f"$/{denominator}"


def extract_effective_date(text: str) -> Optional[str]:
    """Extract an explicitly labelled effective date as ISO-8601."""
    match = re.search(
        r"(?:effective|rates? (?:as of|from))\s*:?[ ]*"
        r"([A-Z][a-z]+\s+\d{1,2},?\s+\d{4}|\d{4}[-/]\d{1,2}[-/]\d{1,2})",
        text,
        re.I,
    )
    if not match:
        return None
    raw = match.group(1).replace("/", "-")
    for fmt in ("%B %d, %Y", "%B %d %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def clean_number(text: str) -> Optional[float]:
    """Parse a numeric string, stripping commas and whitespace."""
    if not text:
        return None
    text = text.strip().replace(",", "").replace(" ", "")
    match = re.search(r"-?\d+\.?\d*", text)
    if match:
        return float(match.group(0))
    return None


def normalize_province(text: str) -> str:
    """
    Normalize a province name to its standard 2-letter code.
    Accepts full names, common abbreviations, and 2-letter codes.
    """
    mapping = {
        "british columbia": "BC", "bc": "BC",
        "alberta": "AB", "ab": "AB",
        "saskatchewan": "SK", "sk": "SK",
        "manitoba": "MB", "mb": "MB",
        "ontario": "ON", "on": "ON",
        "quebec": "QC", "québec": "QC", "qc": "QC",
        "new brunswick": "NB", "nb": "NB",
        "nova scotia": "NS", "ns": "NS",
        "prince edward island": "PE", "pei": "PE", "pe": "PE",
        "newfoundland and labrador": "NL", "newfoundland": "NL", "nl": "NL",
        "yukon": "YT", "yt": "YT",
        "northwest territories": "NT", "nwt": "NT", "nt": "NT",
        "nunavut": "NU", "nu": "NU",
    }
    return mapping.get(text.strip().lower(), text.strip().upper())


# ─── Live-scraping helpers ───────────────────────────────────

def find_text_near_label(
    soup: BeautifulSoup,
    label_text: str,
    search_radius: int = 3,
) -> Optional[str]:
    """
    Find a label (like "Basic Charge") in the page and return numeric
    text found near it. Searches next siblings, parent's next siblings,
    and adjacent table cells within *search_radius* hops.

    Returns the first string that looks like it contains a number,
    or None if nothing is found.
    """
    label_lower = label_text.lower()

    # Search all text nodes containing the label
    for tag in soup.find_all(string=re.compile(re.escape(label_lower), re.IGNORECASE)):
        element = tag.parent if tag.parent else tag

        # Strategy 1: next siblings of the element
        candidates = []
        sib = element.next_sibling
        for _ in range(search_radius):
            if sib is None:
                break
            text = sib.get_text(strip=True) if hasattr(sib, "get_text") else str(sib).strip()
            if text:
                candidates.append(text)
            sib = sib.next_sibling

        # Strategy 2: if element is inside a <td>/<th>, check the next <td>/<th> in the row
        cell = element.find_parent(["td", "th"])
        if cell:
            for next_cell in cell.find_next_siblings(["td", "th"])[:search_radius]:
                text = next_cell.get_text(strip=True)
                if text:
                    candidates.append(text)

        # Strategy 3: parent's next sibling (common in <dt>/<dd> or <div> layouts)
        parent = element.parent
        if parent:
            psib = parent.next_sibling
            for _ in range(search_radius):
                if psib is None:
                    break
                text = psib.get_text(strip=True) if hasattr(psib, "get_text") else str(psib).strip()
                if text:
                    candidates.append(text)
                psib = psib.next_sibling

        # Return the first candidate that contains a digit
        for cand in candidates:
            if re.search(r"\d", cand):
                return cand

    return None


def extract_rate_from_text(text: str) -> Optional[float]:
    """
    Extract a rate value from free-form text. Handles patterns like:
      "$0.0945/kWh"  "12.34 ¢/kWh"  "$25.00/month"  "0.0945 $/GJ"
      "9.45 cents per kWh"  "$6.89 per kW"

    Always returns the value in dollars (converts cents automatically).
    Returns None if no numeric rate is found.
    """
    if not text:
        return None

    text = text.strip()

    # Detect cents
    is_cents = bool(re.search(r"[¢]|cents?\b", text, re.IGNORECASE))

    # Try to extract a dollar amount like $X.XXXX
    match = re.search(r"\$\s*(\d+\.?\d*)", text)
    if match:
        return float(match.group(1))

    # Try bare number (possibly with cents indicator)
    match = re.search(r"(\d+\.?\d*)\s*(?:[¢]|cents?|/|\$|per\b)", text, re.IGNORECASE)
    if not match:
        # Last resort: any number in the text
        match = re.search(r"(\d+\.?\d*)", text)

    if not match:
        return None

    value = float(match.group(1))
    if is_cents:
        value /= 100.0

    return value


def detect_js_rendered(html: str) -> bool:
    """
    Heuristic check for JS-rendered pages. Returns True if the page
    appears to rely on client-side JavaScript for content rendering.

    Indicators:
      - Very little visible text in <body> relative to page size
      - Many <script> tags
      - References to common SPA frameworks (React, Angular, Vue, Next.js)
      - Presence of div#root, div#app, or div#__next with no content
    """
    soup = parse_html(html)
    body = soup.find("body")
    if not body:
        return False

    body_text = body.get_text(strip=True)
    scripts = body.find_all("script")

    # Very little text but lots of scripts
    if len(body_text) < 200 and len(scripts) > 3:
        return True

    # Low text-to-HTML ratio
    html_len = len(html)
    if html_len > 5000 and len(body_text) < html_len * 0.05:
        return True

    # SPA framework markers
    spa_markers = [
        "__NEXT_DATA__", "__next", "react-root", "_react",
        "ng-app", "ng-version", "Vue.js", "nuxt",
    ]
    html_lower = html.lower()
    if any(marker.lower() in html_lower for marker in spa_markers):
        # Check if body has substantive content despite framework markers
        if len(body_text) < 500:
            return True

    # Empty root containers
    for div_id in ["root", "app", "__next", "main-content"]:
        div = body.find("div", id=div_id)
        if div and len(div.get_text(strip=True)) < 50:
            return True

    return False


def find_pdf_links(
    soup: BeautifulSoup,
    keywords: Optional[list[str]] = None,
    base_url: Optional[str] = None,
) -> list[str]:
    """
    Extract all href values linking to PDF files from a parsed page.

    Args:
        soup: BeautifulSoup object of the parsed page.
        keywords: Optional list of keywords to filter by. If provided,
                  only links where the href or link text contains at
                  least one keyword (case-insensitive) are returned.
        base_url: Page URL used to resolve relative PDF links.

    Returns:
        List of PDF URLs (may be relative paths).
    """
    pdf_links = []
    for a_tag in soup.find_all("a", href=True):
        href = a_tag["href"]
        # Some publishers append cache-busting or download query strings.
        if not urlsplit(href).path.lower().endswith(".pdf"):
            continue

        if keywords:
            link_text = a_tag.get_text(strip=True).lower()
            href_lower = href.lower()
            if not any(kw.lower() in link_text or kw.lower() in href_lower for kw in keywords):
                continue

        resolved = urljoin(base_url, href) if base_url else href
        if resolved not in pdf_links:
            pdf_links.append(resolved)

    return pdf_links


def rate_value_appears(
    text: str,
    value: float,
    unit: str,
    *,
    label: Optional[str] = None,
    context_characters: int = 180,
) -> bool:
    """Return whether an expected charge appears in extracted source text.

    This is deliberately strict: it accepts the normal dollar and cent
    renderings of a value, but does not use fuzzy matching.  Scrapers use it
    to prove every fallback component is still present in an official source
    before labelling the resulting tariff as live-verified.
    """
    if not text or value is None:
        return False

    normalized = re.sub(r"\s+", " ", text).replace(",", "")
    dollar_values = {f"{value:.2f}", f"{value:.3f}", f"{value:.4f}", f"{value:.5f}"}
    patterns = []

    if unit == "$/kWh":
        cents = value * 100
        cent_values = {f"{cents:g}", f"{cents:.2f}", f"{cents:.3f}", f"{cents:.4f}"}
        patterns.extend(
            rf"(?<!\d){re.escape(number)}\s*(?:¢|cents?)\s*(?:/|per)\s*kWh"
            for number in cent_values
        )

    unit_name = {
        "$/kWh": r"kWh",
        "$/kW": r"kW",
        "$/kVA": r"kVA",
        "$/month": r"(?:month|mo(?:nthly)?)",
        "$/day": r"day|daily",
        "$/GJ": r"GJ",
        "$/m3": r"m(?:3|³)",
    }.get(unit)
    if unit_name:
        patterns.extend(
            rf"\$\s*{re.escape(number)}\s*(?:/|per)\s*{unit_name}"
            for number in dollar_values
        )

    matches = [m for pattern in patterns for m in re.finditer(pattern, normalized, re.IGNORECASE)]
    if not matches:
        return False
    if not label:
        return True
    label_tokens = [t for t in re.findall(r"[a-z]{3,}", label.casefold()) if t not in {"charge", "rate"}]
    if not label_tokens:
        return False
    for match in matches:
        context = normalized[max(0, match.start() - context_characters):match.end() + context_characters].casefold()
        if any(token in context for token in label_tokens):
            return True
    return False


def verify_tariff_values(
    text: str,
    records: list,
    *,
    require_context: bool = False,
) -> list[str]:
    """List components not proven in their tariff/label/unit context.

    Each tariff's name or code must identify a section when contextual mode is
    enabled. This prevents an unrelated occurrence of the same number elsewhere
    in a large schedule from verifying a component.
    """
    missing = []
    for record in records:
        section = text
        if require_context:
            identifiers = [x for x in (record.tariff_code, record.tariff_name) if x]
            locations = [text.casefold().find(str(x).casefold()) for x in identifiers]
            locations = [loc for loc in locations if loc >= 0]
            if not locations:
                missing.extend(f"{record.tariff_name}: {c.component_name} (tariff section absent)" for c in record.components)
                continue
            start = min(locations)
            section = text[max(0, start - 300):start + 8000]
        for component in record.components:
            if component.charge_value is None or not component.charge_unit:
                continue
            if not rate_value_appears(
                section,
                component.charge_value,
                component.charge_unit,
                label=component.component_name if require_context else None,
            ):
                missing.append(f"{record.tariff_name}: {component.component_name}")
    return missing
